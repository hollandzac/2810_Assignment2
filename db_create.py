import openpyxl as xl
import sqlite3 as sql

connection = sql.connect("compliance.db")
cursor = connection.cursor()


#Creates table for inspection results
cursor.execute("""CREATE TABLE IF NOT EXISTS inspections (
activity_date TEXT,
employee_id VARCHAR(9),
facility_address VARCHAR(40),
facility_city VARCHAR(20),
facility_id VARCHAR(9),
facility_name VARCHAR(30),
facility_state VARCHAR(2),
facility_zip VARCHAR(12),
grade CHAR(1),
owner_id VARCHAR(9),
owner_name VARCHAR(30),
pe_description VARCHAR(50),
program_element_pe INTEGER(4),
program_name VARCHAR(30),
program_status VARCHAR(8),
record_id VARCHAR(9),
score INTEGER(3),
serial_number VARCHAR(9),
service_code INTEGER(3),
service_description VARCHAR(30)
)""")

#Creates table for violation results
cursor.execute("""CREATE TABLE IF NOT EXISTS violations(
points INTEGER(2),
serial_number VARCHAR(9),
violation_code VARCHAR(5),
violation_description VARCHAR(50),
violation_satus VARCHAR(25)
)""")

#Loads the 2 sheets
insp = xl.load_workbook("inspections.xlsx")['inspections']
viol = xl.load_workbook("violations.xlsx")['violations']


#Command for inserting inspection excel into databse
insert = """INSERT INTO inspections VALUES (date('{}'),'{}','{}','{}','{}','{}','{}','{}',
'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')"""

#For each row in the excel sheet creates a list of all cell values the inserts them into command and executes
for row in insp.iter_rows(min_row=2):
	#Replaces single qoutes with double single for sql proccessing
    values = [str(row[i].value).replace("'","''") for i in range(20)]
    values[0] = values[0].strip(' 00:00:00')
    if len(values[0]) < 10:
        values[0] = values[0][:8] + '0' + values[0][8:]
    print(values[0])
    command = insert.format(*values)
    cursor.execute(command)

#command for innsertinf violations excel sheet into database
insert = """INSERT INTO violations VALUES ('{}','{}','{}','{}','{}')"""

#For each row in the excel sheet creates a list of all cell values the inserts them into command and executes
for row in viol.iter_rows(min_row=2):
    values2 = [str(row[i].value).replace("'","''") for i in range(5)]
    command = insert.format(*values2)
    cursor.execute(command)


connection.commit()
connection.close()
