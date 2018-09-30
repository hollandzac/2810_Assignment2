import openpyxl as xl
from openpyxl import Workbook
import sqlite3 as sql

connection = sql.connect("compliance.db")
cursor = connection.cursor()

wb = Workbook()

ws1 = wb.active

ws1.title = 'Violation_Types'
ws1['A1'] = 'Code'
ws1['B1'] = 'Description'
ws1['C1'] = 'Count'


print(wb.sheetnames)

query = """SELECT violation_code, violation_description, COUNT(violation_code)
FROM violations
GROUP BY violation_code ORDER BY violation_code
"""
cursor.execute(query)
result = cursor.fetchall()


for a,b in zip(ws1.iter_rows(min_row=2, min_col=1, max_col=3, max_row=200), result):
    for i in range(3):
        a[i].value = b[i]



wb.save('ViolationTypes.xlsx')
